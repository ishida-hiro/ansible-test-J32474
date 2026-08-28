# -*- coding: utf-8 -*-
"""Regenerate docs/02_ロール一覧.xlsx and docs/03_要確認事項.xlsx from their Markdown masters."""
import re, io, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill('solid', fgColor='DCE6F1')
TITLE_F  = Font(bold=True, size=14, color='1F3864')
HDR_F    = Font(bold=True)
THIN     = Side(style='thin', color='808080')
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP     = Alignment(wrap_text=True, vertical='top')

def strip_md(t):
    t = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', t)
    t = t.replace('**', '').replace('`', '')
    return t.strip()

def md_tables(text):
    """Yield (header, rows) for each pipe table."""
    lines, i = text.split('\n'), 0
    while i < len(lines):
        if re.match(r'^\s*\|.*\|\s*$', lines[i]) and i+1 < len(lines) and re.match(r'^\s*\|[\s:|-]+\|\s*$', lines[i+1]):
            cells = lambda s: [strip_md(c) for c in s.strip().strip('|').split('|')]
            hdr = cells(lines[i]); i += 2; rows = []
            while i < len(lines) and re.match(r'^\s*\|.*\|\s*$', lines[i]):
                rows.append(cells(lines[i])); i += 1
            yield hdr, rows
        else:
            i += 1

def finish(ws, widths, header_row):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    for row in ws.iter_rows(min_row=header_row):
        for c in row:
            if c.value is not None:
                c.alignment = WRAP
                c.border = BORDER
    for c in ws[header_row]:
        if c.value is not None:
            c.font = HDR_F
            c.fill = HDR_FILL
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

# ---------------------------------------------------------------- 02 ロール一覧
def build_02(src, dst):
    text = io.open(src, encoding='utf-8').read()
    tables = list(md_tables(text))
    wb = openpyxl.Workbook()

    ws = wb.active; ws.title = 'ロール一覧と設計書対応表'
    ws['A1'] = 'ロール一覧と設計書対応表'; ws['A1'].font = TITLE_F
    ws['A2'] = '対象設計書: sample_Windows詳細設計書.xlsx（シート「1.概要」「8.ファイルサーバ」）'
    hdr, rows = tables[0]
    ws.append([]) ; ws.append(hdr)
    for r in rows:
        ws.append(r)
    finish(ws, [18, 34, 64, 40], 4)

    ws = wb.create_sheet('変数の定義場所')
    ws['A1'] = '変数の定義場所'; ws['A1'].font = TITLE_F
    hdr, rows = tables[1]
    ws.append([]); ws.append(hdr)
    for r in rows:
        ws.append(r)
    finish(ws, [42, 26, 60], 3)

    ws = wb.create_sheet('冪等性について')
    ws['A1'] = '冪等性について'; ws['A1'].font = TITLE_F
    ws['A1'].alignment = Alignment(vertical='top')
    body = text.split('## 冪等性について', 1)[1]
    r = 3
    for ln in body.split('\n'):
        s = strip_md(ln)
        if not s:
            continue
        ws.cell(row=r, column=1, value=('・' + s[2:] if ln.strip().startswith('- ') else s))
        ws.cell(row=r, column=1).alignment = WRAP
        r += 1
    ws.column_dimensions['A'].width = 110
    wb.save(dst); print('wrote', dst)

# ------------------------------------------------------------ 03 要確認事項
def build_03(src, dst):
    text = io.open(src, encoding='utf-8').read()
    wb = openpyxl.Workbook()

    # --- parse "### A-1. title" sections -------------------------------------
    items = []
    parts = re.split(r'\n### ', text)
    for p in parts[1:]:
        head, _, body = p.partition('\n')
        m = re.match(r'^([ABC])-(\d+)\.\s*(.*)$', head.strip())
        if not m:
            continue
        kubun, no, title = m.group(1), f'{m.group(1)}-{m.group(2)}', strip_md(m.group(3))
        body = body.split('\n---')[0]
        # 「今回の扱い」: 【今回…】マーカー優先。無ければ「現状の実装:」の一文を使う
        marks = [m for m in re.findall(r'【([^】]+)】', body) if m.startswith('今回')]
        seen, handling = set(), []
        for m in marks:
            if m not in seen:
                seen.add(m); handling.append(m)
        handling = ' / '.join(handling)
        if not handling:
            m = re.search(r'\*\*現状の実装\*\*[:：]\s*(.+?)(?:\n\s*\n|$)', body, re.S)
            if m:
                handling = re.sub(r'\s*\n\s*', '', strip_md(m.group(1)))[:120]
        asks = [strip_md(x) for x in re.findall(r'\*\*→\s*(.+?)\*\*', body, re.S)]
        content = []
        for para in re.split(r'\n\s*\n', body):
            para = para.strip()
            if not para or para.startswith('|') or para.startswith('**→'):
                continue
            content.append(re.sub(r'\s*\n\s*', '', strip_md(para)))
        items.append([kubun, no, title, handling,
                      '\n'.join(content[:4]), '\n'.join('→ ' + a for a in asks)])

    ws = wb.active; ws.title = '要確認事項一覧'
    ws['A1'] = '要確認事項 一覧'; ws['A1'].font = TITLE_F
    ws['A2'] = 'A=実行前に必須 / B=実装判断あり（認識合わせが必要） / C=参考・留意事項'
    ws['A3'] = '対象設計書: sample_Windows詳細設計書.xlsx（シート「1.概要」「8.ファイルサーバ」）'
    ws.append([])
    ws.append(['区分', 'No.', '件名', '今回の扱い', '内容', 'ご確認いただきたいこと'])
    for it in items:
        ws.append(it)
    finish(ws, [7, 8, 40, 30, 78, 52], 5)
    ws.auto_filter.ref = f'A5:F{4 + len(items)}'

    # --- 凡例 ---------------------------------------------------------------
    ws = wb.create_sheet('凡例・区分について')
    ws['A1'] = '凡例・区分について'; ws['A1'].font = TITLE_F
    ws.append([]); ws.append(['区分', '意味'])
    for k, v in [('A（実行前に必須）', '値が確定しないと構築できない、または誤った設定になる'),
                 ('B（実装判断あり）', '設計書の記載から解釈して実装した。認識合わせが必要'),
                 ('C（参考）', '動作に影響しうる留意事項')]:
        ws.append([k, v])
    finish(ws, [24, 80], 3)

    # --- 対応状況サマリ ------------------------------------------------------
    tables = list(md_tables(text))
    ws = wb.create_sheet('対応状況サマリ')
    ws['A1'] = '対応状況サマリ（今回の Ansible 初期構築）'; ws['A1'].font = TITLE_F
    hdr, rows = tables[0]
    ws.append([]); ws.append(hdr)
    for r in rows:
        ws.append(r)
    finish(ws, [40, 80], 3)

    # --- 区分ごとのシート ----------------------------------------------------
    for kubun, name in [('A', 'A. 実行前に必須'), ('B', 'B. 実装判断あり'), ('C', 'C. 参考・留意事項')]:
        ws = wb.create_sheet(name)
        ws['A1'] = name; ws['A1'].font = TITLE_F
        ws.append([])
        ws.append(['No.', '件名', '今回の扱い', '内容', 'ご確認いただきたいこと'])
        for it in items:
            if it[0] == kubun:
                ws.append(it[1:])
        finish(ws, [8, 40, 30, 82, 54], 3)

    wb.save(dst); print('wrote', dst, f'({len(items)} items)')

build_02('docs/02_ロール一覧.md', 'docs/02_ロール一覧.xlsx')
build_03('docs/03_要確認事項.md', 'docs/03_要確認事項.xlsx')
